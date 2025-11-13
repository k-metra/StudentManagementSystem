import json
from openpyxl import load_workbook
from termcolor import colored 
from utils.misc import enter_to_continue

def import_students_from_excel(file_path: str) -> dict[str, dict]:

    students = {}

    try:
        workbook = load_workbook(filename=file_path)

        sheet = workbook.active # First sheet by default

        headers = [str(cell.value).strip() for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
        required_fields = {
            'student_id', 'first_name', 'last_name', 'year_level', 'phone_number', 'course', 'home_address',
            'email_address', 'guardian_name', 'guardian_contact'
        }

        data = json.load(open("src/data/departments.json", "r"))
        departments = data.get("departments", {})

        missing = required_fields - set(headers)

        if missing:
            print(colored(f"Error: Missing required columns: {', '.join(missing)}", "red"))
            enter_to_continue()
            return {}
        
        header_indices = {header: idx for idx, header in enumerate(headers)}

        for row in sheet.iter_rows(min_row=2, values_only=True):
            student_id = str(row[header_indices.get('student_id', '')] or '').strip()

            if not student_id:
                continue

            row_department = None

            for department, courses in departments.items():
                if str(row[header_indices.get('course', '')] or '').strip() in courses:
                    row_department = department
                    break
            
            students[student_id] = {
                'first_name': str(row[header_indices.get('first_name', '')] or '').strip(),
                'last_name': str(row[header_indices.get('last_name', '')] or '').strip(),
                'year_level': str(row[header_indices.get('year_level', '')] or "1").strip(),
                'phone_number': str(row[header_indices.get('phone_number', '')] or '').strip(),
                'course': str(row[header_indices.get('course', '')] or '').strip(),
                'home_address': str(row[header_indices.get('home_address', '')] or '').strip(),
                'email_address': str(row[header_indices.get('email_address', '')] or '').strip(),
                'guardian_name': str(row[header_indices.get('guardian_name', '')] or '').strip(),
                'guardian_contact': str(row[header_indices.get('guardian_contact', '')] or '').strip(),
                'department': row_department or 'Undeclared',
            }

        print(colored(f"Successfully imported {len(students)} student records from '{file_path}'.", "green"))
        return students
    
    except FileNotFoundError:
        print(colored(f"Error: The file '{file_path}' was not found.", "red"))
        enter_to_continue()
    except Exception as e:
        print(colored(f"An error occurred while importing: {e}", "red"))

    return {}
           